import openpyxl

from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.styles.borders import Border, Side

ad_review_week = input("Which ad week are you reviewing? ")
filepath = "/Users/Kyle McNamara/Desktop/Work/Code/Gatekeeper/Gatekeeper_Wk"+ad_review_week+"_Working_File.xlsx"
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "GM Wk"+ad_review_week
wrkbk_align = ws1
ws1_align = Alignment(horizontal='center',vertical='bottom')
wrkbk_align.alignment = ws1_align
static_header_font = styles.Font(size=11, bold=True)
static_header_fill = styles.PatternFill(start_color='A9A9A9',end_color='A9A9A9', fill_type='solid')
dynamic_cells_font = styles.Font(size=11, bold=True,color='FFFFFF')
dynamic_cells_alignment = Alignment(textRotation=90,horizontal='center',vertical='bottom')
LRS = styles.PatternFill(start_color='E57A00',end_color='E57A00',fill_type='solid')
RDFB = styles.PatternFill(start_color='00117D',end_color='00117D',fill_type='solid')
Lift = styles.PatternFill(start_color='3F3F40',end_color='3F3F40',fill_type='solid') 
Billed = styles.PatternFill(start_color='292B37',end_color='292B37',fill_type='solid') 
Cuts = styles.PatternFill(start_color='E4B547',end_color='E4B547',fill_type='solid') 
Distros = styles.PatternFill(start_color='D1D1D1',end_color='D1D1D1',fill_type='solid') 
Sales = styles.PatternFill(start_color='9A9A9A',end_color='9A9A9A',fill_type='solid') 
Billed_Sales = styles.PatternFill(start_color='2BA3D7',end_color='2BA3D7',fill_type='solid')
Revised_Booking = styles.PatternFill(start_color='2C8D53',end_color='2C8D53',fill_type='solid')
GK_Booking = styles.PatternFill(start_color='8D2C2C',end_color='8D2C2C',fill_type='solid')
header_border = Border(bottom=Side(border_style='thin',color='000000'))
column_border = Border(left=Side(border_style='thin',color='000000'),
					   right=Side(border_style='thin',color='000000'),
					   bottom=Side(border_style='thin',color='000000'))




format = ['Shipper Key', 'History Key', 'Sales Key', 'FA', 'Desk', 'Dept.' 'Brand', 
		  'Building', 'Product Line', 'Pricing Group', 'Description', 'AD Code', 'AD Sub-Code', 'C&S Code', 
		  'UPC', 'Booked', 'LRS', 'RDFB', 'Lift', 'Billed', 'Cuts', 
		  'Distros', 'Sales', 'Billed/Sales', 'Revised Booking', 'GK Booking', 'Max Billed', 'Max Sales', 
		  'Comments','Ad Type Name', 'Retail Factor', 'Retail Amount', 'Copient', 'Memo', 'Ad Type', 
		  'Start Date', 'End Date', 'Planned Distro', 'Lead Time', 'Replenishment Analyst', 'Merchandiser', 'Manufacturer', 
		  'Manufacturer-Lo']

for col, val in enumerate(format, start=1):
	ws1.cell(row=1, column=col).value = val

for cell in ws1["1:1"]:
	cell.font = static_header_font
	cell.fill = static_header_fill
	cell.border = header_border
	
ws1['P1'].fill=LRS
ws1['Q1'].fill=RDFB
ws1['R1'].fill=Lift
ws1['S1'].fill=Billed
ws1['T1'].fill=Cuts
ws1['U1'].fill=Distros
ws1['V1'].fill=Sales
ws1['W1'].fill=Billed_Sales
ws1['X1'].fill=Revised_Booking
ws1['Y1'].fill=GK_Booking
ws1['P1'].font=dynamic_cells_font
ws1['Q1'].font=dynamic_cells_font
ws1['R1'].font=dynamic_cells_font
ws1['S1'].font=dynamic_cells_font
ws1['T1'].font=dynamic_cells_font
ws1['U1'].font=dynamic_cells_font
ws1['V1'].font=dynamic_cells_font
ws1['W1'].font=dynamic_cells_font
ws1['X1'].font=dynamic_cells_font
ws1['Y1'].font=dynamic_cells_font

	
for row in ws1:
	for cell in row:
		cell.alignment = ws1_align

ws1['P1'].alignment=dynamic_cells_alignment
ws1['Q1'].alignment=dynamic_cells_alignment
ws1['R1'].alignment=dynamic_cells_alignment
ws1['S1'].alignment=dynamic_cells_alignment
ws1['T1'].alignment=dynamic_cells_alignment
ws1['U1'].alignment=dynamic_cells_alignment
ws1['V1'].alignment=dynamic_cells_alignment
ws1['W1'].alignment=dynamic_cells_alignment
ws1['X1'].alignment=dynamic_cells_alignment
ws1['Y1'].alignment=dynamic_cells_alignment
ws1['Q1'].border=column_border
ws1['Z1'].border=column_border
ws1['AA1'].border=column_border
	
wb.save(filepath)


	
	
