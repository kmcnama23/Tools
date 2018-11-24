import xlsxwriter
import openpyxl
import openpyxl as xl
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import coordinate_from_string, column_index_from_string
from shutil import copyfile
from copy import copy


ad_review_week = input("Which ad week are you reviewing? ")
Gatekeeper = xlsxwriter.Workbook("C:/Users/Kyle McNamara/Desktop/Work/Code/Gatekeeper/Gatekeeper_Wk"+ad_review_week+"_Working_File.xlsx") 
gk = Gatekeeper.add_worksheet("GM Wk"+ad_review_week)
hist = Gatekeeper.add_worksheet("GM History")
sales = Gatekeeper.add_worksheet("GM Sales")
Gatekeeper.close()
source_path = 'C:\\Users\\Kyle McNamara\\Desktop\\Work\\Gatekeeper_Information\\Week '+ad_review_week+', 2019\\GM_Wk'+ad_review_week+'_GK_File.xlsx'
destination_path = 'C:\\Users\\Kyle McNamara\\Desktop\\Work\\Code\\Gatekeeper\\Gatekeeper_Wk'+ad_review_week+'_Working_File.xlsx'
hist_sales_path ='C:\\Users\\Kyle McNamara\\Desktop\\Work\\Gatekeeper_Information\\Week '+ad_review_week+', 2019\\GM_History_&_Sales.xlsx'	
gkdata = xl.load_workbook(filename=source_path)
wrkngfile=xl.load_workbook(filename=destination_path)
histdata = xl.load_workbook(filename=hist_sales_path)
sourcedata = gkdata["GM Wk"+ad_review_week]
billdata=histdata["GM History"]
salesdata=histdata["GM Sales"]
gktarget=wrkngfile["GM Wk"+ad_review_week]
histtarget=wrkngfile["GM History"]
salestarget=wrkngfile["GM Sales"] 
sheet=gkdata["GM Wk"+ad_review_week]
active_rows = gkdata.worksheets[0]
LRS = styles.PatternFill(start_color='E57A00',end_color='E57A00',fill_type='solid')
RDFB = styles.PatternFill(start_color='00117D',end_color='00117D',fill_type='solid')
Lift = styles.PatternFill(start_color='3F3F40',end_color='3F3F40',fill_type='solid') 
Billed = styles.PatternFill(start_color='292B37',end_color='292B37',fill_type='solid') 
Cuts = styles.PatternFill(start_color='E4B547',end_color='E4B547',fill_type='solid') 
Distros = styles.PatternFill(start_color='D1D1D1',end_color='D1D1D1',fill_type='solid') 
Sales = styles.PatternFill(start_color='9A9A9A',end_color='9A9A9A',fill_type='solid') 
Billed_Sales = styles.PatternFill(start_color='2BA3D7',end_color='2BA3D7',fill_type='solid')
Revised_Booking = styles.PatternFill(start_color='2C8D53',end_color='2C8D53',fill_type='solid')
GK_Booking = styles.PatternFill(start_color='8D2C2C',end_color='8D2C2C',fill_type='solid')
static_header_fill = styles.PatternFill(start_color='A9A9A9',end_color='A9A9A9', fill_type='solid')
static_header_font = styles.Font(size=11, bold=True, color='000000')
static_alignment = Alignment(horizontal='center',vertical='bottom')
dynamic_cells_font = styles.Font(size=11, bold=True,color='FFFFFF')
dynamic_cells_alignment = Alignment(textRotation=90,horizontal='center',vertical='bottom')
header_border = Border(bottom=Side(border_style='thin',color='000000'))
column_border = Border(left=Side(border_style='thin',color='000000'),
						right=Side(border_style='thin',color='000000'),
						bottom=Side(border_style='thin',color='000000'))
wrkbk_align = gktarget
gktarget_align = Alignment(horizontal='center',vertical='bottom')
wrkbk_align.alignment = gktarget_align
row_count = active_rows.max_row
col_count = active_rows.max_column

class GkImport(object):
	
	
	def __init__(self, startCol, startRow, endCol, endRow, sheet):
		self.startCol=startCol
		self.startRow=startRow
		self.endCol=endCol
		self.endRow=endRow
		
	def gkCopy(self):
		startCol=1
		startRow=1
		endCol=col_count
		endRow=row_count
		rangeSelected = []
		#Loops through selected Rows
		for i in range(startRow,endRow + 1,1):
		#Appends the row to a RowSelected list
			rowSelected = []
			for j in range(startCol,endCol+1,1):
				rowSelected.append(sheet.cell(row = i, column = j).value)
			#Adds the RowSelected List and nests inside the rangeSelected
			rangeSelected.append(rowSelected)
		return rangeSelected
	
ingest = GkImport(1,1,col_count,row_count,sourcedata)

def paste(firstCol, firstRow, lastCol, lastRow, sheetReceiving,copiedData):
	numberRow = 0
	for m in range(firstRow,lastRow+1,1):
		numberCol = 0
		for o in range(firstCol,lastCol+1,1):
		
			sheetReceiving.cell(row = m , column = o).value = copiedData[numberRow][numberCol]
			numberCol += 1
		numberRow += 1
		
def input():
	gkContent = ingest.gkCopy()
	destination=paste(1,1,col_count,row_count,gktarget,gkContent)
input()		

class HistImport(object):
	
	
	def __init__(self, firstCol, firstRow, lastCol, lastRow, secondSheet):
		self.startCol=firstCol
		self.startRow=firstRow
		self.endCol=lastCol
		self.endRow=lastRow
		
	def histCopy(self):
		firstCol=1
		firstRow=1
		lastCol=col_count
		lastRow=row_count
		contentSelected = []
		secondSheet=billdata
		#Loops through selected Rows
		for a in range(firstRow,lastRow + 1,1):
		#Appends the row to a RowSelected list
			secondrowSelected = []
			for b in range(firstCol,lastCol+1,1):
				secondrowSelected.append(secondSheet.cell(row = a, column = b).value)
			#Adds the RowSelected List and nests inside the rangeSelected
			contentSelected.append(secondrowSelected)
		return contentSelected
				
digest = HistImport(1,1,col_count,row_count,billdata)

def histPaste(firstCol, firstRow, lastCol, lastRow, sheetReceiving, copiedData):
	numberRow = 0
	for m in range(firstRow,lastRow+1,1):
		numberCol = 0
		for o in range(firstCol,lastCol+1,1):
		
			sheetReceiving.cell(row = m , column = o).value = copiedData[numberRow][numberCol]
			numberCol += 1
		numberRow += 1
		
def input1():
	histContent = digest.histCopy()
	destination=paste(1,1,col_count,row_count,histtarget,histContent)
input1()	

class SalesImport(object):
	
	
	def __init__(self, beginCol, beginRow, termCol, termRow, thirdSheet):
		self.beginCol=beginCol
		self.beginRow=beginRow
		self.termCol=termCol
		self.termRow=termRow
		
	def salesCopy(self):
		beginCol=1
		beginRow=1
		termCol=col_count
		termRow=row_count
		dataSelected = []
		thirdSheet=salesdata
		#Loops through selected Rows
		for c in range(beginRow,termRow + 1,1):
		#Appends the row to a RowSelected list
			thirdrowSelected = []
			for d in range(beginCol,termCol+1,1):
				thirdrowSelected.append(thirdSheet.cell(row = c, column = d).value)
			#Adds the RowSelected List and nests inside the rangeSelected
			dataSelected.append(thirdrowSelected)
		return dataSelected

absorb = SalesImport(1,1,col_count,row_count,salesdata)

def salesPaste(firstCol, firstRow, lastCol, lastRow, sheetReceiving, copiedData):
	numberRow = 0
	for m in range(firstRow,lastRow+1,1):
		numberCol = 0
		for o in range(firstCol,lastCol+1,1):
		
			sheetReceiving.cell(row = m , column = o).value = copiedData[numberRow][numberCol]
			numberCol += 1
		numberRow += 1
		
def input2():
	salesContent = absorb.salesCopy()
	destination=paste(1,1,col_count,row_count,salestarget,salesContent)
input2()

gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)
gktarget.insert_cols(19)

format = ['Lift', 'Billed', 'Cuts', 'Distros', 'Sales', 'Billed/Sales', 'Revised Booking', 
		'GK Booking', 'Max Billed', 'Max Sales']
			
#Column Writer
for col, val in enumerate(format, start=19):
	gktarget.cell(row=1, column=col).value = val


gktarget['A1'].fill=static_header_fill
gktarget['B1'].fill=static_header_fill
gktarget['C1'].fill=static_header_fill
gktarget['D1'].fill=static_header_fill
gktarget['E1'].fill=static_header_fill
gktarget['F1'].fill=static_header_fill
gktarget['G1'].fill=static_header_fill
gktarget['H1'].fill=static_header_fill
gktarget['I1'].fill=static_header_fill
gktarget['J1'].fill=static_header_fill
gktarget['K1'].fill=static_header_fill
gktarget['L1'].fill=static_header_fill
gktarget['M1'].fill=static_header_fill
gktarget['N1'].fill=static_header_fill
gktarget['O1'].fill=static_header_fill
gktarget['P1'].fill=static_header_fill
gktarget['AD1'].fill=static_header_fill
gktarget['AE1'].fill=static_header_fill
gktarget['AF1'].fill=static_header_fill
gktarget['AG1'].fill=static_header_fill
gktarget['AH1'].fill=static_header_fill
gktarget['AI1'].fill=static_header_fill
gktarget['AJ1'].fill=static_header_fill
gktarget['AK1'].fill=static_header_fill
gktarget['AL1'].fill=static_header_fill
gktarget['AM1'].fill=static_header_fill
gktarget['AN1'].fill=static_header_fill
gktarget['AO1'].fill=static_header_fill
gktarget['AP1'].fill=static_header_fill
gktarget['AQ1'].fill=static_header_fill
gktarget['AA1'].fill=static_header_fill
gktarget['AB1'].fill=static_header_fill
gktarget['AC1'].fill=static_header_fill
gktarget['A1'].font=static_header_font
gktarget['B1'].font=static_header_font
gktarget['C1'].font=static_header_font
gktarget['D1'].font=static_header_font
gktarget['E1'].font=static_header_font
gktarget['F1'].font=static_header_font
gktarget['G1'].font=static_header_font
gktarget['H1'].font=static_header_font
gktarget['I1'].font=static_header_font
gktarget['J1'].font=static_header_font
gktarget['K1'].font=static_header_font
gktarget['L1'].font=static_header_font
gktarget['M1'].font=static_header_font
gktarget['N1'].font=static_header_font
gktarget['O1'].font=static_header_font
gktarget['P1'].font=static_header_font
gktarget['AD1'].font=static_header_font
gktarget['AE1'].font=static_header_font
gktarget['AF1'].font=static_header_font
gktarget['AG1'].font=static_header_font
gktarget['AH1'].font=static_header_font
gktarget['AI1'].font=static_header_font
gktarget['AJ1'].font=static_header_font
gktarget['AK1'].font=static_header_font
gktarget['AL1'].font=static_header_font
gktarget['AM1'].font=static_header_font
gktarget['AN1'].font=static_header_font
gktarget['AO1'].font=static_header_font
gktarget['AP1'].font=static_header_font
gktarget['AQ1'].font=static_header_font
gktarget['AA1'].font=static_header_font
gktarget['AB1'].font=static_header_font
gktarget['AC1'].font=static_header_font
gktarget['A1'].border=column_border
gktarget['AA1'].border=column_border
gktarget['AB1'].border=column_border

###Alignment
for row in gktarget:
	for cell in row:
		cell.alignment = gktarget_align	

###Dynamic Formatting
gktarget['Q1'].fill=LRS
gktarget['R1'].fill=RDFB
gktarget['S1'].fill=Lift
gktarget['T1'].fill=Billed
gktarget['U1'].fill=Cuts
gktarget['V1'].fill=Distros
gktarget['W1'].fill=Sales
gktarget['X1'].fill=Billed_Sales
gktarget['Y1'].fill=Revised_Booking
gktarget['Z1'].fill=GK_Booking
gktarget['Q1'].font=dynamic_cells_font
gktarget['R1'].font=dynamic_cells_font
gktarget['S1'].font=dynamic_cells_font
gktarget['T1'].font=dynamic_cells_font
gktarget['U1'].font=dynamic_cells_font
gktarget['V1'].font=dynamic_cells_font
gktarget['W1'].font=dynamic_cells_font
gktarget['X1'].font=dynamic_cells_font
gktarget['Y1'].font=dynamic_cells_font
gktarget['Z1'].font=dynamic_cells_font
gktarget['Q1'].alignment=dynamic_cells_alignment
gktarget['R1'].alignment=dynamic_cells_alignment
gktarget['S1'].alignment=dynamic_cells_alignment
gktarget['T1'].alignment=dynamic_cells_alignment
gktarget['U1'].alignment=dynamic_cells_alignment
gktarget['V1'].alignment=dynamic_cells_alignment
gktarget['W1'].alignment=dynamic_cells_alignment
gktarget['X1'].alignment=dynamic_cells_alignment
gktarget['Y1'].alignment=dynamic_cells_alignment
gktarget['Z1'].alignment=dynamic_cells_alignment
	
for col in gktarget.columns:
	max_length = 0
	column = col[0].column
	for cell in col:
		try:
			if len(str(cell.value)) > max_length:
				max_length = len(cell.value)
		except:
			pass
	adjusted_width = (max_length + 2) * 1.1 
	gktarget.column_dimensions[column].width = adjusted_width

for col in histtarget.columns:
	max_length = 0
	column = col[0].column
	for cell in col:
		try:
			if len(str(cell.value)) > max_length:
				max_length = len(cell.value)
		except:
			pass
	adjusted_width = (max_length + 2) * 1.1 
	histtarget.column_dimensions[column].width = adjusted_width

for col in salestarget.columns:
	max_length = 0
	column = col[0].column
	for cell in col:
		try:
			if len(str(cell.value)) > max_length:
				max_length = len(cell.value)
		except:
			pass
	adjusted_width = (max_length + 2) * 1.1 
	salestarget.column_dimensions[column].width = adjusted_width	






























	

wrkngfile.save(destination_path)





		