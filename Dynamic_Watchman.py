import openpyxl
import openpyxl as xl
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import coordinate_from_string, column_index_from_string
from shutil import copyfile

ad_review_week = input("Which ad week are you reviewing?(again) ")	
source_path='C:\\Users\\Kyle McNamara\\Desktop\\Work\\Gatekeeper_Information\\Week '+ad_review_week+', 2019\\GM_Wk'+ad_review_week+'_GK_File.xlsx'
destination_path='C:\\Users\\Kyle McNamara\\Desktop\\Work\\Code\\Gatekeeper\\Gatekeeper_Wk'+ad_review_week+'_Working_File.xlsx'
hist_sales_path ='C:\\Users\\Kyle McNamara\\Desktop\\Work\\Gatekeeper_Information\\Week '+ad_review_week+', 2019\\GM_Wk'+ad_review_week+'_GK_File.xlsx'
##Loads copy-from source file/selects first worksheet
wb1 = xl.load_workbook(filename=source_path)
ws1 = wb1.worksheets[0]
working_keeper = "Working Keeper"

#Load copy-to source file/creates duplicate worksheet
wb2 = xl.load_workbook(filename=destination_path)
ws2 = wb2.create_sheet()
ws2.title = 'Sheet2'
	
for row in ws1:
	for cell in row:
		ws2[cell.coordinate].value = cell.value
		

wb2.save(destination_path)



wb = openpyxl.load_workbook(destination_path)
RSC137 = wb.active
wrkbk_align = RSC137
RSC137_align = Alignment(horizontal='center',vertical='bottom')
wrkbk_align.alignment = RSC137_align
LRS = styles.PatternFill(start_color='E57A00',end_color='E57A00',fill_type='solid')
RDFB = styles.PatternFill(start_color='00117D',end_color='00117D',fill_type='solid')
Lift = styles.PatternFill(start_color='3F3F40',end_color='3F3F40',fill_type='solid') 
Billed = styles.PatternFill(start_color='292B37',end_color='292B37',fill_type='solid') 
Cuts = styles.PatternFill(start_color='E4B547',end_color='E4B547',fill_type='solid') 
Distros = styles.PatternFill(start_color='D1D1D1',end_color='D1D1D1',fill_type='solid') 
Sales = styles.PatternFill(start_color='9A9A9A',end_color='9A9A9A',fill_type='solid') 
Billed_Sales = styles.PatternFill(start_color='2BA3D7',end_color='2BA3D7',fill_type='solid')
Revised_Booking = styles.PatternFill(start_color='2C8D53',end_color='2C8D53',fill_type='solid')
GK_Booking = styles.PatternFill(start_color='8D2C2C',end_color='8D2C2C',fill_type='solid')
static_header_fill = styles.PatternFill(start_color='A9A9A9',end_color='A9A9A9', fill_type='solid')
static_header_font = styles.Font(size=11, bold=True, color='000000')
static_alignment = Alignment(horizontal='center',vertical='bottom')
dynamic_cells_font = styles.Font(size=11, bold=True,color='FFFFFF')
dynamic_cells_alignment = Alignment(textRotation=90,horizontal='center',vertical='bottom')
header_border = Border(bottom=Side(border_style='thin',color='000000'))
column_border = Border(left=Side(border_style='thin',color='000000'),
						right=Side(border_style='thin',color='000000'),
						bottom=Side(border_style='thin',color='000000'))
row_count = ws2.max_row
target = wb["Sheet1"]
source = wb["Sheet2"]
fomula_value = RSC137.cell(row=2,column=19)
	
#Insert columns - pretty straight forward	
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)
RSC137.insert_cols(19)

		
format = ['Lift', 'Billed', 'Cuts', 'Distros', 'Sales', 'Billed/Sales', 'Revised Booking', 
		'GK Booking', 'Max Billed', 'Max Sales']
			
#Column Writer
for col, val in enumerate(format, start=19):
	RSC137.cell(row=1, column=col).value = val
		

def copyRange(startCol, startRow, endCol, endRow, sheet):
	rangeSelected = []
	#Loops through selected Rows
	for i in range(startRow,endRow + 1,1):
	#Appends the row to a RowSelected list
		rowSelected = []
		for j in range(startCol,endCol+1,1):
			rowSelected.append(sheet.cell(row = i, column = j).value)
		#Adds the RowSelected List and nests inside the rangeSelected
		rangeSelected.append(rowSelected)
 
	return rangeSelected
         
 
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):

	countRow = 0
	for i in range(startRow,endRow+1,1):
		countCol = 0
		for j in range(startCol,endCol+1,1):
            
			sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
			countCol += 1
		countRow += 1
		
def createFirst():
	print("Processing...")
	selectedRange = copyRange(1,1,18,row_count,source) 
	pastingRange = pasteRange(1,1,18,row_count,target,selectedRange) 
	wb.save(destination_path)
	

def secondCopy(firstCol, firstRow, lastCol, lastRow, sheet):
	contentSelected=[]
	for m in range(firstRow,lastRow+1,1):
		cellsSelected=[]
		for o in range(firstCol,lastCol+1,1):
			cellsSelected.append(sheet.cell(row=m,column=o).value)
		contentSelected.append(cellsSelected)
	
	return contentSelected

def secondPaste(firstCol, firstRow, lastCol, lastRow, sheetReceiving, data):
	numberRow = 0
	for m in range(firstRow,lastRow+1,1):
		numberCol = 0
		for o in range(firstCol,lastCol+1,1):
		
			sheetReceiving.cell(row = m , column = o).value = data[numberRow][numberCol]
			numberCol += 1
		numberRow += 1
		
	
def createSecond():
	selectedContent = secondCopy(19,1,33,row_count,source) 
	pastingCells = secondPaste(29,1,43,row_count,target,selectedContent)
	wb.save(destination_path)
		

createFirst()
createSecond()

#Static Cell Formatting
RSC137['A1'].fill=static_header_fill
RSC137['B1'].fill=static_header_fill
RSC137['C1'].fill=static_header_fill
RSC137['D1'].fill=static_header_fill
RSC137['E1'].fill=static_header_fill
RSC137['F1'].fill=static_header_fill
RSC137['G1'].fill=static_header_fill
RSC137['H1'].fill=static_header_fill
RSC137['I1'].fill=static_header_fill
RSC137['J1'].fill=static_header_fill
RSC137['K1'].fill=static_header_fill
RSC137['L1'].fill=static_header_fill
RSC137['M1'].fill=static_header_fill
RSC137['N1'].fill=static_header_fill
RSC137['O1'].fill=static_header_fill
RSC137['P1'].fill=static_header_fill
RSC137['AD1'].fill=static_header_fill
RSC137['AE1'].fill=static_header_fill
RSC137['AF1'].fill=static_header_fill
RSC137['AG1'].fill=static_header_fill
RSC137['AH1'].fill=static_header_fill
RSC137['AI1'].fill=static_header_fill
RSC137['AJ1'].fill=static_header_fill
RSC137['AK1'].fill=static_header_fill
RSC137['AL1'].fill=static_header_fill
RSC137['AM1'].fill=static_header_fill
RSC137['AN1'].fill=static_header_fill
RSC137['AO1'].fill=static_header_fill
RSC137['AP1'].fill=static_header_fill
RSC137['AQ1'].fill=static_header_fill
RSC137['AA1'].fill=static_header_fill
RSC137['AB1'].fill=static_header_fill
RSC137['AC1'].fill=static_header_fill
RSC137['A1'].font=static_header_font
RSC137['B1'].font=static_header_font
RSC137['C1'].font=static_header_font
RSC137['D1'].font=static_header_font
RSC137['E1'].font=static_header_font
RSC137['F1'].font=static_header_font
RSC137['G1'].font=static_header_font
RSC137['H1'].font=static_header_font
RSC137['I1'].font=static_header_font
RSC137['J1'].font=static_header_font
RSC137['K1'].font=static_header_font
RSC137['L1'].font=static_header_font
RSC137['M1'].font=static_header_font
RSC137['N1'].font=static_header_font
RSC137['O1'].font=static_header_font
RSC137['P1'].font=static_header_font
RSC137['AD1'].font=static_header_font
RSC137['AE1'].font=static_header_font
RSC137['AF1'].font=static_header_font
RSC137['AG1'].font=static_header_font
RSC137['AH1'].font=static_header_font
RSC137['AI1'].font=static_header_font
RSC137['AJ1'].font=static_header_font
RSC137['AK1'].font=static_header_font
RSC137['AL1'].font=static_header_font
RSC137['AM1'].font=static_header_font
RSC137['AN1'].font=static_header_font
RSC137['AO1'].font=static_header_font
RSC137['AP1'].font=static_header_font
RSC137['AQ1'].font=static_header_font
RSC137['AA1'].font=static_header_font
RSC137['AB1'].font=static_header_font
RSC137['AC1'].font=static_header_font
RSC137['A1'].border=column_border
RSC137['AA1'].border=column_border
RSC137['AB1'].border=column_border



###Alignment
for row in RSC137:
	for cell in row:
		cell.alignment = RSC137_align	

###Dynamic Formatting
RSC137['Q1'].fill=LRS
RSC137['R1'].fill=RDFB
RSC137['S1'].fill=Lift
RSC137['T1'].fill=Billed
RSC137['U1'].fill=Cuts
RSC137['V1'].fill=Distros
RSC137['W1'].fill=Sales
RSC137['X1'].fill=Billed_Sales
RSC137['Y1'].fill=Revised_Booking
RSC137['Z1'].fill=GK_Booking
RSC137['Q1'].font=dynamic_cells_font
RSC137['R1'].font=dynamic_cells_font
RSC137['S1'].font=dynamic_cells_font
RSC137['T1'].font=dynamic_cells_font
RSC137['U1'].font=dynamic_cells_font
RSC137['V1'].font=dynamic_cells_font
RSC137['W1'].font=dynamic_cells_font
RSC137['X1'].font=dynamic_cells_font
RSC137['Y1'].font=dynamic_cells_font
RSC137['Z1'].font=dynamic_cells_font
RSC137['Q1'].alignment=dynamic_cells_alignment
RSC137['R1'].alignment=dynamic_cells_alignment
RSC137['S1'].alignment=dynamic_cells_alignment
RSC137['T1'].alignment=dynamic_cells_alignment
RSC137['U1'].alignment=dynamic_cells_alignment
RSC137['V1'].alignment=dynamic_cells_alignment
RSC137['W1'].alignment=dynamic_cells_alignment
RSC137['X1'].alignment=dynamic_cells_alignment
RSC137['Y1'].alignment=dynamic_cells_alignment
RSC137['Z1'].alignment=dynamic_cells_alignment

remove_sht = wb['Sheet2']
wb.remove(remove_sht)
		
###Autofit
for col in RSC137.columns:
	max_length = 0
	column = col[0].column
	for cell in col:
		try:
			if len(str(cell.value)) > max_length:
				max_length = len(cell.value)
		except:
			pass
	adjusted_width = (max_length + 2) * 1.1 
	RSC137.column_dimensions[column].width = adjusted_width			

RSC137.title = "GM Wk"+ad_review_week


	
wb.save(destination_path)

print("Data Input Successful. Please execute Venerable_Watchman.py to begin billed and sales history data import.")
