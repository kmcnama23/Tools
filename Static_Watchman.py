import openpyxl
import openpyxl as xl
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import coordinate_from_string, column_index_from_string
from shutil import copyfile
import xlsxwriter 


class Static_Watchman():
	ad_review_week = input("Which ad week are you reviewing? ")
	copyfile("/Users/Kyle McNamara/Desktop/Work/Gatekeeper_Information/Week "+ad_review_week+", 2019/GM_Wk"+ad_review_week+"_GK_File.xlsx",
			"/Users/Kyle McNamara/Desktop/Work/Code/Gatekeeper/Gatekeeper_Wk"+ad_review_week+"_Working_File.xlsx")
	filepath = "/Users/Kyle McNamara/Desktop/Work/Code/Gatekeeper/Gatekeeper_Wk"+ad_review_week+"_Working_File.xlsx"
	wb = openpyxl.load_workbook(filepath)
	ws1 = wb.active
	ws2 = wb.create_sheet("GM History")
	ws3 = wb.create_sheet("GM Sales")
	ws4 = wb.create_sheet("Beth4 Wk"+ad_review_week)
	ws5 = wb.create_sheet("Beth4 History")
	ws6 = wb.create_sheet("Beth4 Sales")
	ws7 = wb.create_sheet("MI")
	ws1.title = "GM Wk"+ad_review_week
	wrkbk_align = ws1
	ws1_align = Alignment(horizontal='center',vertical='bottom')
	wrkbk_align.alignment = ws1_align
	dynamic_cells_font = styles.Font(size=11, bold=True,color='FFFFFF')
	dynamic_cells_alignment = Alignment(textRotation=90,horizontal='center',vertical='bottom')
	static_header_font = styles.Font(size=11, bold=True, color='000000')
	static_header_fill = styles.PatternFill(start_color='A9A9A9',end_color='A9A9A9', fill_type='solid')
	LRS = styles.PatternFill(start_color='E57A00',end_color='E57A00',fill_type='solid')
	RDFB = styles.PatternFill(start_color='00117D',end_color='00117D',fill_type='solid')
	Lift = styles.PatternFill(start_color='3F3F40',end_color='3F3F40',fill_type='solid') 
	Billed = styles.PatternFill(start_color='292B37',end_color='292B37',fill_type='solid') 
	Cuts = styles.PatternFill(start_color='E4B547',end_color='E4B547',fill_type='solid') 
	Distros = styles.PatternFill(start_color='D1D1D1',end_color='D1D1D1',fill_type='solid') 
	Sales = styles.PatternFill(start_color='9A9A9A',end_color='9A9A9A',fill_type='solid') 
	Billed_Sales = styles.PatternFill(start_color='2BA3D7',end_color='2BA3D7',fill_type='solid')
	Revised_Booking = styles.PatternFill(start_color='2C8D53',end_color='2C8D53',fill_type='solid')
	GK_Booking = styles.PatternFill(start_color='8D2C2C',end_color='8D2C2C',fill_type='solid')
	header_border = Border(bottom=Side(border_style='thin',color='000000'))
	column_border = Border(left=Side(border_style='thin',color='000000'),
						   right=Side(border_style='thin',color='000000'),
						   bottom=Side(border_style='thin',color='000000'))
		
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	ws1.insert_cols(19)
	
	
	format = ['Lift', 'Billed', 'Cuts', 'Distros', 'Sales', 'Billed/Sales', 'Revised Booking', 
			'GK Booking', 'Max Billed', 'Max Sales']
			
	###Column Writer
	for col, val in enumerate(format, start=19):
		ws1.cell(row=1, column=col).value = val
		
	for row in ws1:
		for cell in row:
			cell.alignment = ws1_align
	
	#Dynamic Cell Formatting
	ws1['Q1'].fill=LRS
	ws1['R1'].fill=RDFB
	ws1['S1'].fill=Lift
	ws1['T1'].fill=Billed
	ws1['U1'].fill=Cuts
	ws1['V1'].fill=Distros
	ws1['W1'].fill=Sales
	ws1['X1'].fill=Billed_Sales
	ws1['Y1'].fill=Revised_Booking
	ws1['Z1'].fill=GK_Booking
	ws1['AA1'].fill=static_header_fill
	ws1['AB1'].fill=static_header_fill
	ws1['AC1'].fill=static_header_fill
	ws1['Q1'].font=dynamic_cells_font
	ws1['R1'].font=dynamic_cells_font
	ws1['S1'].font=dynamic_cells_font
	ws1['T1'].font=dynamic_cells_font
	ws1['U1'].font=dynamic_cells_font
	ws1['V1'].font=dynamic_cells_font
	ws1['W1'].font=dynamic_cells_font
	ws1['X1'].font=dynamic_cells_font
	ws1['Y1'].font=dynamic_cells_font
	ws1['Z1'].font=dynamic_cells_font
	ws1['AA1'].font=static_header_font
	ws1['AB1'].font=static_header_font
	ws1['AC1'].font=static_header_font
	ws1['Q1'].alignment=dynamic_cells_alignment
	ws1['R1'].alignment=dynamic_cells_alignment
	ws1['S1'].alignment=dynamic_cells_alignment
	ws1['T1'].alignment=dynamic_cells_alignment
	ws1['U1'].alignment=dynamic_cells_alignment
	ws1['V1'].alignment=dynamic_cells_alignment
	ws1['W1'].alignment=dynamic_cells_alignment
	ws1['X1'].alignment=dynamic_cells_alignment
	ws1['Y1'].alignment=dynamic_cells_alignment
	ws1['Z1'].alignment=dynamic_cells_alignment
	ws1['A1'].border=column_border
	ws1['AA1'].border=column_border
	ws1['AB1'].border=column_border		
	
	
	s = 1
	x = '$T${0}/$W${0}'
	row_count = ws1.max_row
	row = 1
	column_r = 18
	column_s = 19
	column_x = 24
	increment = 1
	
	while row < row_count:
		#ws1.cell(row=row+increment,column=column_r).value = "ROUND($S${0}*($T${0}+$U${0}*.6)/5,0)*5}"
		ws1.cell(column=column_s,row=row+increment,value=s)
		row += 1
	print(column_r)
	print(row)
	
	###Autofit
	for col in ws1.columns:
		max_length = 0
		column = col[0].column
		for cell in col:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws1.column_dimensions[column].width = adjusted_width	
	wb.save(filepath)